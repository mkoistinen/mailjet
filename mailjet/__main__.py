# -*- coding: utf8 -*-

import argparse
import binascii
import json
import os
import socket
from typing import Iterable, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import phonenumbers
from phonenumbers import PhoneNumberFormat as Formats
from phonenumbers.phonenumberutil import NumberParseException
import requests

from .settings import ACCESS_TOKEN


API_URL = 'https://api.mailjet.com/v4/sms-send'

SMS_SIZES = [
    (160, 70), (306, 134), (459, 201), (612, 268), (765, 335)
]

GSM7_CHARSET = (
    "@£$¥èéùìòÇ\nØø\rÅåΔ_ΦΓΛΩΠΨΣΘΞ\x1bÆæßÉ !\"#¤%&'()*+,-./0123456789:;<=>?"  # noqa
    "¡ABCDEFGHIJKLMNOPQRSTUVWXYZÄÖÑÜ§¿abcdefghijklmnopqrstuvwxyzäöñüà"   # noqa
)
EXT_CHARSET = (
    "````````````````````^```````````````````{}`````\\````````````[~]`"
    "|````````````````````````````````````€``````````````````````````"
)


class RecipientsNotFoundException(Exception):
    """Indicate that a column of recipients could not be found."""


class UnableToCleanException(Exception):
    """Indicate that a recipient could not be cleaned."""


def gsm_encode(plaintext):
    """Convert unicode to GSM-7 encoded text."""
    # Inspired by https://stackoverflow.com/a/2453027
    chars_omitted = []
    res = ""
    for c in plaintext:
        idx = GSM7_CHARSET.find(c)
        if idx != -1:
            res += chr(idx)
            continue
        idx = EXT_CHARSET.find(c)
        if idx != -1:
            res += chr(27) + chr(idx)
        chars_omitted.append(c)
    return chars_omitted, binascii.b2a_hex(res.encode('utf-8'))


def send_sms(msg: str, recipient_list: Iterable[str],
             sender: str = None, dry_run: bool = True) -> None:
    """
    Send a `msg` to a `recipient`.

    Parameters
    ----------
    msg : str
        The message to send
    recipient_list : Iterable[str]
        An iterable of recipient telephone numbers in international format.
    sender : str
        The sender ID
    dry_run : bool
        If set, then "pretend" to send the messages.

    Returns
    -------
    None
    """
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "content-type": "application/json",
    }
    data = {
        "From": sender,
    }

    if isinstance(msg, bytes):
        data['Text'] = f'{msg}'
    else:
        data['Text'] = msg

    for recipient in recipient_list:
        data["To"] = recipient
        if dry_run:
            print(f'DRY_RUN: Pretend a message was sent to {recipient}')
        else:
            res = requests.post(url=API_URL, data=json.dumps(data),
                                headers=headers)
            if res.status_code != 200:
                print(f'Message not sent to {recipient}. An error '
                      f'({res.status_code}) occurred: "{res.content}"')


def find_recipient_data(workbook: Workbook) -> Tuple[Worksheet, int]:
    """
    If found, return the worksheet and column index that contains recipients.

    Do this by scanning the top row for a header named 'sms', 'cell', 'mobile',
    or 'telephone' in any letter case.

    Parameters
    ----------
    workbook : Workbook

    Returns
    -------
    Worksheet
        The worksheet where recipients were found or `None` if none found.
    int
        The column number where recipients were found or `None` if none found.

    Raises
    ------
    RecipientsNotFoundException
        If no appropriate column could be found.
    """
    for sheet in workbook.worksheets:
        for c in range(1, sheet.max_column):
            val = sheet.cell(1, c).value
            if val.lower() in ['sms', 'cell', 'mobile', 'telephone']:
                return sheet, c

    raise RecipientsNotFoundException('Could not find an appropriate column.')


def get_recipients(file_path: str = None) -> Optional[List[str]]:
    """
    Get recipients from the provided file_path.

    Parameters
    ----------
    file_path : str
        The path to the file.

    Returns
    -------
    List[str] or None
        A list of recipients (telephone numbers)
    """
    if not file_path:
        raise ValueError('A valid file path is required.')
    book = load_workbook(filename=file_path, read_only=True)

    try:
        sheet_name, column_num = find_recipient_data(book)
        recipients_list = [
            cell[0] for cell in sheet_name.iter_rows(
                min_row=2, min_col=column_num, max_col=column_num,
                values_only=True)
        ]
    except RecipientsNotFoundException:
        print('An XLSX workbook that contains a worksheet containing a '
              'column of telephone numbers and a header of one of {"sms", '
              '"cell", "mobile", or "telephone"} that contains recipients, '
              'SMS-able phone numbers is required.')
    except Exception:  # noqa: Intentionally broad
        print('An unanticipated error occurred while attempting to extract '
              'the recipients from the XLSX file.')
    else:
        return recipients_list


def clean_phone_numbers(phone_list) -> Set[str]:
    """
    Clean a list of phone numbers into a set of E164 formatted numbers.

    Parameters
    ----------
    phone_list : List[str]

    Returns
    -------
    Set[str]
        A set of the cleaned numbers in E164 format or None if ANY number
        could not be parsed into E.164.

    Raises
    ------
    UnableToCleanException
        If one or more entries were not parse-able into E.164 format.
    """
    error_found = False
    cleaned_numbers = set()

    # Strip out LRO characters that seem prevalent in the sample file.
    phone_list = [p.replace('\u202d', '') for p in phone_list]

    for num in phone_list:
        try:
            parsed = phonenumbers.parse(num)
        except (NumberParseException, TypeError):
            print(f'The telephone number provided: "{num}" raised an '
                  f'unexpected error while attempting to parse it. Please '
                  f'ensure this entry is a valid international (E.164) '
                  f'format-able telephone number and try again.')
            error_found = True
        else:
            if phonenumbers.is_valid_number(parsed):
                cleaned_numbers.add(
                    phonenumbers.format_number(parsed, Formats.E164))
            else:
                print(f'The telephone number provided: "{num}" does not '
                      f'appear to be a valid telephone number. Please correct '
                      f'its format to be in the international (E.164) format '
                      f'(E.g. +12125551212 or +442012341234) and try again.')
                error_found = True

    if error_found:
        raise UnableToCleanException()
    else:
        return cleaned_numbers


def count_messages(num_characters, mode='gsm7') -> Optional[int]:
    """Calculate the number of concatenated messages would be required."""
    for idx, (gsm7, utf16) in enumerate(SMS_SIZES):
        if mode == 'gsm7':
            if gsm7 >= num_characters:
                return idx + 1
        else:
            if utf16 >= num_characters:
                return idx + 1
    return None


def clean_message(message: str, dry_run: bool) -> bool:
    """
    Examine message for sending as SMS. If dry_run, make recommendations.

    Return the message as a GSM7-encoded string, if possible, else as unicode.

    Parameters
    ----------
    message : str
        The message to send
    dry_run : bool
        if set, make recommendations, otherwise only report issues.

    Returns
    -------
    bool
        True if message is send-able.
    """
    characters_omitted, gsm7_msg = gsm_encode(message)
    num_gsm7_msgs = count_messages(len(message), 'gsm7')
    num_utf16_msgs = count_messages(len(message), 'utf16')

    if characters_omitted:
        if num_utf16_msgs is None:
            # This is a Unicode message and it is too big to send. ERROR.
            print(f'ERROR: This message is too long to be sent via SMS in '
                  f'UTF-16 format.')
            if dry_run and num_gsm7_msgs is not None:
                print(f'However, if the symbols {characters_omitted} were '
                      f'replaced with GSM-7 friendly symbols, it could be '
                      f'sent as a GSM-7 message.')
            return False
        elif dry_run:
            savings = (num_utf16_msgs - num_gsm7_msgs) / num_utf16_msgs
            print(f'SUGGESTION: If the following symbols {characters_omitted} '
                  f'are replaced with GSM7 encode-able symbols, the message '
                  f'would be sent as a GSM7-encoded message and would cost '
                  f'~{savings * 100.0:0.0f}% less to send.')
        mode = 'utf16'

    else:
        mode = 'gsm7'

    if mode == 'gsm7':
        if num_gsm7_msgs and num_gsm7_msgs > 1:
            print(f'This message will be sent as {num_gsm7_msgs} '
                  f'concatenated GSM-7 encoded messages.')
        else:
            print('This message will be sent as a single GSM-7 '
                  'encoded message.')
    else:
        if num_utf16_msgs and num_utf16_msgs > 1:
            print(f'This message will be sent as {num_utf16_msgs} '
                  f'concatenated UTF-16 encoded messages.')
        else:
            print('This message will be sent as a single UTF-16 '
                  'encoded message.')
    return True


def run(*, message: str = None,
        recipients: str = None,
        xlsx_file: str = None,
        sender: str = None,
        dry_run: bool = True,
        ) -> None:
    """
    Programmatic entry-point to sending SMS messages via MailJet.com.

    Parameters
    ----------
    message : str
        The message to send
    recipients : Optional[str], default None
        If set, do everything but only send a test-message to the recipient
        provided here.
    xlsx_file : Optional[str]
        A valid path to an XLSX file containing recipients. The spreadsheet
        will be scanned for a column heading in row 0 that contains "SMS",
        "CELL", "PHONE", "MOBILE", or "TELEPHONE" to identify where the
        list of recipients SMS-capable telephone numbers are.
    sender : str, default the `USER` env var
        The sender ID
    dry_run : bool, default True
        If set, do everything EXCEPT make a request to the MailJet API.
    """
    if sender is None:
        sender = os.getenv('USER', socket.gethostname())
    print(f'Using a "sender" value of "{sender}". '
          f'Use the `-s` option to modify.')

    # Examine the message. Note, this doesn't alter the message.
    message_clean = clean_message(message, dry_run)

    # NOTE: Either `recipients` OR `xlsx_file` will be passed (never both)
    if xlsx_file:
        try:
            recipients = get_recipients(file_path=xlsx_file)
            if not recipients:
                print('Could not find any recipients in the provided '
                      'XLSX file.')
        except Exception:  # noqa: Intentionally broad
            print('There were unexpected errors getting the telephone numbers '
                  'from the provided XLSX file.')
            return

    if recipients:
        try:
            recipients = clean_phone_numbers(recipients)
        except UnableToCleanException:
            print('One or more of the the recipient telephone numbers were '
                  'not parsable. Please see the output log to identify the '
                  'problematic entries, correct them, and try again.')
            recipients = None

    if message_clean and recipients:
        send_sms(message, recipient_list=recipients, sender=sender,
                 dry_run=dry_run)
    else:
        print('There is nothing to do.')
        return


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Blast out SMS messages via MailJet.com')
    group = parser.add_mutually_exclusive_group()
    group.add_argument('-r', '--recipients', nargs='*', type=str,
                       help='A single recipient’s phone number. This option '
                             'may be used multiple times if required.')
    group.add_argument('-f', '--xlsx_file', type=str,
                       help='The path to a ".XLSX" file with the a column '
                             'headed with "SMS" containing the target '
                             'phone numbers.')

    parser.add_argument('-m', '--message', type=str,
                        help='The message to send.')
    parser.add_argument_group(group)
    parser.add_argument('-s', '--sender', type=str,
                        help='The sender ID. If not specified, will be the '
                             'value of the ENV var: `USER` if found, or '
                             'the hostname.')
    parser.add_argument('--for-real', dest='dry_run',
                        default=True, action='store_false')

    args = parser.parse_args()
    run(**vars(args))
